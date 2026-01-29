import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from copy import copy
from openpyxl.cell.cell import MergedCell

# ---------------- CONFIG ----------------
SENDER_EMAIL = "jinjutar.smartdev@gmail.com"
SENDER_PASSWORD = "UZFS BDTC XCLZ RZSQ"
RECEIVER_EMAIL = "jinjutar.smartdev@gmail.com"

TEMPLATE_FILE = "template.xlsx"
MAIN_SHEET = "1"
IMAGE_TEMPLATE_SHEET = "ImageTemplate"
# ----------------------------------------


def add_image_to_excel(ws, img_file, cell):
    if img_file is None:
        return

    img_data = io.BytesIO(img_file.getvalue())
    img = Image(img_data)

    # resize ให้พอดี
    max_w, max_h = 600, 350
    ratio = min(max_w / img.width, max_h / img.height)
    img.width = int(img.width * ratio)
    img.height = int(img.height * ratio)

    ws.add_image(img, cell)


def write_safe(ws, cell, value):
    if value is None:
        value = ""

    for merged in ws.merged_cells.ranges:
        if cell in merged:
            ws.cell(row=merged.min_row, column=merged.min_col).value = value
            return

    ws[cell].value = value


# ---------------- UI ----------------
st.set_page_config("Smart Dev Report Generator v0.1", layout="wide")
st.title("🚀 Smart Dev Report Generator v0.1 🚀")

if "photos" not in st.session_state:
    st.session_state.photos = [0]

# --- Part 1 ---
st.subheader("📄 Part 1: Document Details")
c1, c2, c3 = st.columns(3)
doc_no = c1.text_input("Doc. No.")
ref_po = c2.text_input("Ref. PO No.")
date_issue = c3.date_input("Date", datetime.now())

# --- Part 2 ---
st.subheader("🏢 Part 2: Project & Contact Information")
p1, p2 = st.columns(2)

project_name = p1.text_input("Project Name")
site_location = p1.text_input("Site / Location")

contact_client = p2.text_input("Contact Person (Client)")
contact_co_ltd = p2.text_input("Contact (Smart Dev Co., Ltd.)")

engineer_name = st.text_input("Engineer Name (Prepared By)")

# --- Part 3 ---
st.subheader("🛠 Part 3: Service Details")
service_type = st.selectbox(
    "Service Type",
    ["Project", "Commissioning", "Repairing", "Services", "Training", "Check", "Other"]
)

job_performed = st.text_area("Job Performed (รายละเอียดงาน)", height=150)

# --- Part 4 ---
st.subheader("📷 Photo Upload")
final_photo_data = []

for i in list(st.session_state.photos):
    col1, col2, col3 = st.columns([2, 4, 1])

    with col2:
        img = st.file_uploader(f"Image {i+1}", key=f"img{i}", type=["jpg", "png", "jpeg"])
        desc = st.text_input(f"Description {i+1}", key=f"desc{i}")

    with col1:
        if img:
            st.image(img, width=180)

    with col3:
        if st.button("❌", key=f"del{i}"):
            st.session_state.photos.remove(i)
            st.rerun()

    final_photo_data.append({"img": img, "desc": desc})


if st.button("➕ Add Photo"):
    st.session_state.photos.append(max(st.session_state.photos) + 1)
    st.rerun()


# ---------------- GENERATE ----------------
if st.button("🚀 Generate Report"):
    try:
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb[MAIN_SHEET]
        ws_temp = wb[IMAGE_TEMPLATE_SHEET]

        # --- Part 1–3 ---
        write_safe(ws, "B5", doc_no)
        write_safe(ws, "F6", ref_po)
        write_safe(ws, "J5", date_issue.strftime("%d/%m/%Y"))
        write_safe(ws, "B16", project_name)
        write_safe(ws, "H7", site_location)
        write_safe(ws, "C9", contact_client)
        write_safe(ws, "A7", contact_co_ltd)
        write_safe(ws, "B42", engineer_name)
        write_safe(ws, "D15", service_type)
        write_safe(ws, "D17", job_performed)

        # --- รูป 1–6 ---
        img_cells = ["A49","A62","A75","A92","A105","A118"]
        desc_cells = ["H49","H62","H75","H92","H105","H118"]

        for i, item in enumerate(final_photo_data[:6]):
            if item["img"]:
                add_image_to_excel(ws, item["img"], img_cells[i])
                write_safe(ws, desc_cells[i], item["desc"])

        # --- รูป 7+ ต่อท้าย sheet เดียว ---
        extra_photos = final_photo_data[6:]

        if extra_photos:
            block_height = ws_temp.max_row
            start_row = 119

            for page_i in range(0, len(extra_photos), 3):
                group = extra_photos[page_i:page_i+3]

                # copy template
                for r in range(1, ws_temp.max_row + 1):
                    ws.row_dimensions[start_row + r - 1].height = ws_temp.row_dimensions[r].height

                    for c in range(1, ws_temp.max_column + 1):
                        src = ws_temp.cell(r, c)
                        dst = ws.cell(start_row + r - 1, c)

                        if not isinstance(dst, MergedCell):
                            dst.value = src.value
                        if src.has_style:
                            dst.font = copy(src.font)
                            dst.border = copy(src.border)
                            dst.fill = copy(src.fill)
                            dst.number_format = copy(src.number_format)
                            dst.alignment = copy(src.alignment)

                # merged cells
                for m in ws_temp.merged_cells.ranges:
                    new_range = (
                        f"{get_column_letter(m.min_col)}{start_row + m.min_row - 1}:"
                        f"{get_column_letter(m.max_col)}{start_row + m.max_row - 1}"
                    )
                    ws.merge_cells(new_range)

                # ใส่รูป
                img_rows = [5, 18, 31]
                desc_rows = [5, 18, 31]

                for i, item in enumerate(group):
                    if item["img"]:
                        add_image_to_excel(ws, item["img"], f"A{start_row + img_rows[i] - 1}")
                        write_safe(ws, f"H{start_row + desc_rows[i] - 1}", item["desc"])

                start_row += block_height + 1

        # Save output
        output = io.BytesIO()
        wb.save(output)

        # --- EMAIL ---
        msg = MIMEMultipart()
        msg["From"] = SENDER_EMAIL
        msg["To"] = RECEIVER_EMAIL
        msg["Subject"] = f"Report {doc_no}"

        part = MIMEBase("application", "octet-stream")
        part.set_payload(output.getvalue())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename=Report_{doc_no}.xlsx")
        msg.attach(part)

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)

        st.success("✅ Report created and sent!")
                # --- ก้อนโค้ดหมูเด้งเด้งโชว์ตัว ---
        st.balloons()  # ปล่อยลูกโป่งฉลองทั่วจอ
        st.toast("ส่งรายงานสำเร็จแล้วนะหมูเด้ง!", icon='🦛') # เด้งเตือนเล็กๆ มุมขวาล่าง
        
        # แสดงรูปหมูเด้งแบบที่ 2 (ดึงจากเครื่อง)
        col_1, col_2, col_3 = st.columns([1, 2, 1])
        with col_2:
            try:
                # อย่าลืมตั้งชื่อไฟล์รูปในเครื่องให้ตรงกับตรงนี้ (moodeng_ok.png)
                st.image("moodeng_ok.jpg", use_container_width=True) 
                st.markdown("<h3 style='text-align: center; color: #2E7D32;'>เยี่ยมมากหมูเด้ง! ส่งงานเรียบร้อย 🦛✨</h3>", unsafe_allow_html=True)
            except:
                # ถ้าหาไฟล์รูปไม่เจอ ให้โชว์ข้อความสำเร็จแทนเครื่องหมายตกใจ
                st.success("เยี่ยมมาก! ส่งงานเรียบร้อยแล้วจ้า 🦛✨")

        # ปุ่มดาวน์โหลด

        st.download_button(
            "⬇️ Download Report",
            output.getvalue(),
            f"Report_{doc_no}.xlsx"
        )

    except Exception as e:
        st.error(f"❌ Error: {e}")
